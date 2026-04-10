import os
import openpyxl
import datetime
from flask import make_response
import shutil
from dotenv import load_dotenv

# EXCEL出力先
load_dotenv('.env')
DATABASE_URL = os.getenv("DATABASE_URL")
TEMPLATE_FILE = os.getenv("TEMPLATE_FILE")
OUTPUT_DIR = os.getenv("OUTPUT_DIR")

#ユーザー定義フォーマットを追加
currentJPY='"¥"#,##0'

def writeExcel(exportJobList,agencyName,accountName,projectName):
    #今日の日付を取得して発行日として入力
    nowDate=datetime.date.today()

    OUTPUT_FILE = OUTPUT_DIR + str(nowDate) + '-estimate.xlsx'

    shutil.copy(TEMPLATE_FILE,OUTPUT_FILE)

    workbook = openpyxl.load_workbook(OUTPUT_FILE)
    worksheet=workbook.get_sheet_by_name(u'シート1')

    offset = 0
    #各項目と金額を入力
    for i in range(len(exportJobList)):
        if (exportJobList[i][1]):
            offset = 13
            #何月分の作業
            worksheet["A12"].value=exportJobList[i][0]
            #作業内容
            worksheet.cell(row=i+offset,column=1).value=exportJobList[i][1]
        else:
            offset = 12
            #作業内容
            worksheet.cell(row=i+offset,column=1).value=exportJobList[i][0]

        #作業した日付
        worksheet.cell(row=i+offset,column=2).value=exportJobList[i][3]
        #税率
        print(exportJobList[i][4])
        tax_rate = int(exportJobList[i][4] * 100)
        worksheet.cell(row = i + offset, column = 4).value = str(tax_rate) + "%"
        
        #offset = 13
        #作業した日付
        worksheet.cell(row=i+offset,column=2).value=exportJobList[i][3]
        #税率
        #print(exportJobList[i][4])
        tax_rate = int(exportJobList[i][4] * 100)
        worksheet.cell(row = i + offset, column = 4).value = str(tax_rate) + "%"
        #金額
        worksheet.cell(row=i+offset,column=5).number_format=currentJPY
        worksheet.cell(row=i+offset,column=5).value=exportJobList[i][2]
    
    #日付を入力
    worksheet["E3"].value=str(nowDate)
    #代理店名を入力
    worksheet["A3"].value=str(agencyName)
    #担当者名を入力
    worksheet["A4"].value=str(accountName)
    #プロジェクト名を入力
    worksheet["A11"].value=str(projectName)
    #小計を入力
    worksheet["E40"].number_format=currentJPY
    worksheet["E40"].value="=SUM(E12:E39)"

    # 8%税率対象の合計金額
    worksheet["B41"].number_format = currentJPY
    worksheet["B41"].value = '=SUMIF(D12:D39,"8%",E12:E39)'

    #8%税率の税額をC41セルに入力
    worksheet["C41"].number_format = currentJPY
    worksheet["C41"].value = '=B41*0.08'

    #10%税率対象の合計金額
    worksheet["B42"].number_format = currentJPY
    worksheet["B42"].value = '=SUMIF(D12:D39,"10%",E12:E39)'

    #10%税率の税額をC42セルに入力
    worksheet["C42"].number_format = currentJPY
    worksheet["C42"].value = '=B42*0.10'

    #0%税率対象の合計金額
    worksheet["B43"].number_format = currentJPY
    worksheet["B43"].value = '=SUMIF(D12:D39,"0%",E12:E39)'
    
    #0%税率の税額をC43セルに入力
    worksheet["C43"].number_format = currentJPY
    worksheet["C43"].value = '=B43*0'
    
    #消費税合計金額をC44セルに入力
    worksheet["C44"].number_format = currentJPY
    worksheet["C44"].value = "=SUM(C41:C43)"

    #消費税合計金額をE41セルに入力
    worksheet["E41"].number_format = currentJPY
    worksheet["E41"].value = "=SUM(C41:C43)"

    #消費税込み合計金額を入力
    worksheet["E42"].number_format=currentJPY
    worksheet["E42"].value="=SUM(E40:E41)"
    #上にも合計金額を入力
    worksheet["B7"].number_format=currentJPY
    worksheet["B7"].value="=SUM(E40:E41)"
   
    #新しいファイルとして保存する
    workbook.save(OUTPUT_FILE)
    workbook.close()
    
    XLSX_MIMETYPE="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    response=make_response()

    workbook=open(OUTPUT_FILE,"rb")
    response.data=workbook.read()
    workbook.close()

    response.headers["Content-Disposition"]="attachment; filename=" + str(nowDate)+"-estimate.xlsx"
    response.mimetype=XLSX_MIMETYPE
    os.remove(OUTPUT_FILE)
    return response